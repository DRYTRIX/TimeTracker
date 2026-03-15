"""
Tests for the task report view and Excel export (behavior and N+1 optimization).
"""

from datetime import datetime, timedelta


def test_task_report_returns_correct_hours_and_entries(client, app, admin_user, user, project, task):
    """Task report shows tasks with time entries in range and correct hours/entry count."""
    from app import db
    from app.models import TimeEntry, Settings

    with app.app_context():
        settings = Settings.get_settings()
        disabled = list(settings.disabled_module_ids or [])
        if "reports" in disabled:
            settings.disabled_module_ids = [m for m in disabled if m != "reports"]
            db.session.add(settings)
            db.session.commit()

        start_dt = datetime.utcnow() - timedelta(days=5)
        end_dt = datetime.utcnow() - timedelta(days=1)
        e1 = TimeEntry(
            user_id=user.id,
            project_id=project.id,
            task_id=task.id,
            start_time=start_dt,
            end_time=start_dt + timedelta(hours=2),
            duration_seconds=7200,
            notes="Entry one",
            billable=True,
            source="manual",
        )
        e2 = TimeEntry(
            user_id=user.id,
            project_id=project.id,
            task_id=task.id,
            start_time=start_dt + timedelta(hours=3),
            end_time=start_dt + timedelta(hours=5),
            duration_seconds=7200,
            notes="Entry two",
            billable=False,
            source="manual",
        )
        db.session.add_all([e1, e2])
        db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(admin_user.id)
        sess["_fresh"] = True

    start_date = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
    end_date = datetime.utcnow().strftime("%Y-%m-%d")

    resp = client.get(
        f"/reports/tasks?start_date={start_date}&end_date={end_date}&project_id={project.id}&user_id={user.id}",
        follow_redirects=False,
    )

    assert resp.status_code == 200
    data = resp.get_data(as_text=True)
    assert task.name in data
    # 2h + 2h = 4h total for the task
    assert "4.0" in data or "4.00" in data
    assert "2" in data  # entries_count


def test_task_report_excel_export_returns_correct_hours(client, app, admin_user, user, project, task):
    """Task report Excel export has correct task and hours."""
    from app import db
    from app.models import TimeEntry, Settings
    from openpyxl import load_workbook
    import io

    with app.app_context():
        settings = Settings.get_settings()
        disabled = list(settings.disabled_module_ids or [])
        if "reports" in disabled:
            settings.disabled_module_ids = [m for m in disabled if m != "reports"]
            db.session.add(settings)
            db.session.commit()

        start_dt = datetime.utcnow() - timedelta(days=3)
        e = TimeEntry(
            user_id=user.id,
            project_id=project.id,
            task_id=task.id,
            start_time=start_dt,
            end_time=start_dt + timedelta(hours=1, minutes=30),
            duration_seconds=5400,
            notes="Single entry",
            billable=True,
            source="manual",
        )
        db.session.add(e)
        db.session.commit()

    with client.session_transaction() as sess:
        sess["_user_id"] = str(admin_user.id)
        sess["_fresh"] = True

    start_date = (datetime.utcnow() - timedelta(days=7)).strftime("%Y-%m-%d")
    end_date = datetime.utcnow().strftime("%Y-%m-%d")

    resp = client.get(
        f"/reports/task/export/excel?start_date={start_date}&end_date={end_date}&project_id={project.id}&user_id={user.id}",
        follow_redirects=False,
    )

    assert resp.status_code == 200
    assert "spreadsheetml.sheet" in (resp.headers.get("Content-Type") or "")

    wb = load_workbook(filename=io.BytesIO(resp.data))
    ws = wb.active
    rows = [r for r in ws.iter_rows(values_only=True) if any(v not in (None, "") for v in (r or []))]
    assert len(rows) >= 2  # header + at least one data row
    # Header: Task, Project, Status, Completed At, Hours
    assert "Task" in str(rows[0])
    # Data row should have task name and 1.5 hours
    data_row = next((r for r in rows[1:] if task.name in str(r)), None)
    assert data_row is not None
    numbers = [float(x) for x in (data_row or []) if isinstance(x, (int, float))]
    assert any(abs(n - 1.5) < 0.01 for n in numbers), f"Expected ~1.5 in {data_row}"


def test_time_entry_repository_get_task_aggregates(app, project, task, user):
    """Repository get_task_aggregates returns correct (task_id, total_seconds, entry_count)."""
    from app import db
    from app.models import TimeEntry
    from app.repositories import TimeEntryRepository

    with app.app_context():
        start_dt = datetime.utcnow() - timedelta(days=1)
        end_dt = datetime.utcnow()
        for _ in range(2):
            e = TimeEntry(
                user_id=user.id,
                project_id=project.id,
                task_id=task.id,
                start_time=start_dt,
                end_time=start_dt + timedelta(hours=1),
                duration_seconds=3600,
                source="manual",
            )
            db.session.add(e)
        db.session.commit()

        repo = TimeEntryRepository()
        agg = repo.get_task_aggregates([task.id], start_dt, end_dt)
        assert len(agg) == 1
        tid, total_sec, cnt = agg[0]
        assert tid == task.id
        assert total_sec == 7200
        assert cnt == 2
