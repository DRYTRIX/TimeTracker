import io
from datetime import datetime, timedelta


def _non_empty_rows(ws):
    rows = []
    for row in ws.iter_rows(values_only=True):
        if any(v not in (None, "") for v in row):
            rows.append(row)
    return rows


def test_export_user_entries_excel_returns_one_row_per_entry(client, app, admin_user, user, project, task):
    from app import db
    from app.models import TimeEntry, Settings
    from openpyxl import load_workbook

    with app.app_context():
        # Ensure the reports module is enabled for the test (module_enabled decorator).
        settings = Settings.get_settings()
        disabled = list(settings.disabled_module_ids or [])
        if "reports" in disabled:
            settings.disabled_module_ids = [m for m in disabled if m != "reports"]
            db.session.add(settings)
            db.session.commit()

        now = datetime.utcnow()
        start_dt = now - timedelta(hours=3)
        end_dt1 = now - timedelta(hours=2)
        end_dt2 = now - timedelta(hours=1)

        e1 = TimeEntry(
            user_id=user.id,
            project_id=project.id,
            task_id=task.id,
            start_time=start_dt,
            end_time=end_dt1,
            notes="Entry one",
            billable=True,
            source="manual",
        )
        e2 = TimeEntry(
            user_id=user.id,
            project_id=project.id,
            task_id=task.id,
            start_time=end_dt1,
            end_time=end_dt2,
            notes="Entry two",
            billable=False,
            source="manual",
        )
        db.session.add_all([e1, e2])
        db.session.commit()

    # Authenticate as admin via session (avoids host/cookie edge cases in tests)
    with client.session_transaction() as sess:
        sess["_user_id"] = str(admin_user.id)
        sess["_fresh"] = True

    start_date = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    end_date = datetime.utcnow().strftime("%Y-%m-%d")

    resp = client.get(
        f"/reports/user/export/entries/excel?"
        f"user_id={user.id}&project_id={project.id}&start_date={start_date}&end_date={end_date}"
        f"&columns=date&columns=user&columns=project&columns=task&columns=duration_hours&columns=notes",
        follow_redirects=False,
    )

    assert resp.status_code == 200, f"Unexpected redirect to {resp.headers.get('Location')}"
    assert (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        in (resp.headers.get("Content-Type") or "")
    )
    assert ".xlsx" in (resp.headers.get("Content-Disposition") or "")

    wb = load_workbook(filename=io.BytesIO(resp.data))
    ws = wb.active

    rows = _non_empty_rows(ws)
    assert len(rows) == 1 + 2  # header + 2 entries

    header = list(rows[0])
    assert header == ["Date", "User", "Project", "Task", "Duration (hours)", "Notes"]

    # Notes should match the inserted entries (order is newest-first)
    notes = [rows[1][5], rows[2][5]]
    assert set(notes) == {"Entry one", "Entry two"}


def test_export_user_entries_excel_non_admin_cannot_export_other_users(authenticated_client, app, admin_user):
    start_date = (datetime.utcnow() - timedelta(days=1)).strftime("%Y-%m-%d")
    end_date = datetime.utcnow().strftime("%Y-%m-%d")

    resp = authenticated_client.get(
        f"/reports/user/export/entries/excel?user_id={admin_user.id}&start_date={start_date}&end_date={end_date}",
        follow_redirects=False,
    )

    assert resp.status_code in (302, 303)
