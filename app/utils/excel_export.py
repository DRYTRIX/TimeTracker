"""Excel export utilities for reports and data export"""

import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def create_time_entries_excel(entries, filename_prefix='timetracker_export'):
    """Create Excel file from time entries
    
    Args:
        entries: List of TimeEntry objects
        filename_prefix: Prefix for the filename
        
    Returns:
        tuple: (BytesIO object with Excel file, filename)
    """
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Time Entries"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'ID', 'User', 'Project', 'Client', 'Task', 'Start Time', 'End Time',
        'Duration (hours)', 'Duration (formatted)', 'Notes', 'Tags',
        'Source', 'Billable', 'Created At'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, entry in enumerate(entries, 2):
        data = [
            entry.id,
            entry.user.display_name if entry.user else 'Unknown',
            entry.project.name if entry.project else 'N/A',
            entry.project.client.name if (entry.project and entry.project.client) else 'N/A',
            entry.task.name if entry.task else 'N/A',
            entry.start_time.isoformat() if entry.start_time else '',
            entry.end_time.isoformat() if entry.end_time else '',
            entry.duration_hours if entry.end_time else 0,
            entry.duration_formatted if entry.end_time else 'In Progress',
            entry.notes or '',
            entry.tags or '',
            entry.source or 'manual',
            'Yes' if entry.billable else 'No',
            entry.created_at.isoformat() if entry.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Format duration column as number
            if col_num == 8 and isinstance(value, (int, float)):
                cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary at the bottom
    last_row = len(entries) + 2
    ws.cell(row=last_row + 1, column=1, value="Summary")
    ws.cell(row=last_row + 1, column=1).font = Font(bold=True)
    
    total_hours = sum(e.duration_hours for e in entries if e.end_time)
    billable_hours = sum(e.duration_hours for e in entries if e.end_time and e.billable)
    
    ws.cell(row=last_row + 2, column=1, value="Total Hours:")
    ws.cell(row=last_row + 2, column=2, value=total_hours).number_format = '0.00'
    ws.cell(row=last_row + 3, column=1, value="Billable Hours:")
    ws.cell(row=last_row + 3, column=2, value=billable_hours).number_format = '0.00'
    ws.cell(row=last_row + 4, column=1, value="Total Entries:")
    ws.cell(row=last_row + 4, column=2, value=len(entries))
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'{filename_prefix}_{timestamp}.xlsx'
    
    return output, filename


def create_project_report_excel(projects_data, start_date, end_date):
    """Create Excel file for project report
    
    Args:
        projects_data: List of project dictionaries with hours and costs
        start_date: Report start date
        end_date: Report end date
        
    Returns:
        tuple: (BytesIO object with Excel file, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Report"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add report header
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f"Project Report: {start_date} to {end_date}"
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center")
    
    # Column headers
    headers = [
        'Project', 'Client', 'Total Hours', 'Billable Hours',
        'Hourly Rate', 'Billable Amount', 'Total Costs', 'Total Value'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
    
    # Write project data
    for row_num, project in enumerate(projects_data, 4):
        data = [
            project.get('name', ''),
            project.get('client', ''),
            project.get('total_hours', 0),
            project.get('billable_hours', 0),
            project.get('hourly_rate', 0),
            project.get('billable_amount', 0),
            project.get('total_costs', 0),
            project.get('total_value', 0),
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Format numbers
            if col_num in [3, 4]:  # Hours
                cell.number_format = '0.00'
            elif col_num in [5, 6, 7, 8]:  # Money
                cell.number_format = '#,##0.00'
    
    # Auto-adjust columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 40)
        ws.column_dimensions[column].width = adjusted_width
    
    # Add totals
    last_row = len(projects_data) + 4
    ws.cell(row=last_row + 1, column=1, value="TOTALS").font = Font(bold=True)
    
    total_hours = sum(p.get('total_hours', 0) for p in projects_data)
    total_billable_hours = sum(p.get('billable_hours', 0) for p in projects_data)
    total_amount = sum(p.get('billable_amount', 0) for p in projects_data)
    total_costs = sum(p.get('total_costs', 0) for p in projects_data)
    total_value = sum(p.get('total_value', 0) for p in projects_data)
    
    ws.cell(row=last_row + 1, column=3, value=total_hours).number_format = '0.00'
    ws.cell(row=last_row + 1, column=4, value=total_billable_hours).number_format = '0.00'
    ws.cell(row=last_row + 1, column=6, value=total_amount).number_format = '#,##0.00'
    ws.cell(row=last_row + 1, column=7, value=total_costs).number_format = '#,##0.00'
    ws.cell(row=last_row + 1, column=8, value=total_value).number_format = '#,##0.00'
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'project_report_{start_date}_to_{end_date}.xlsx'
    return output, filename


def create_invoice_excel(invoice, items):
    """Create Excel file for a single invoice
    
    Args:
        invoice: Invoice object
        items: List of InvoiceItem objects
        
    Returns:
        tuple: (BytesIO object with Excel file, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    
    # Invoice header
    ws.merge_cells('A1:D1')
    ws['A1'] = f"INVOICE {invoice.invoice_number}"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Invoice details
    ws['A3'] = "Client:"
    ws['B3'] = invoice.client_name
    ws['A4'] = "Issue Date:"
    ws['B4'] = invoice.issue_date.strftime('%Y-%m-%d')
    ws['A5'] = "Due Date:"
    ws['B5'] = invoice.due_date.strftime('%Y-%m-%d')
    ws['A6'] = "Status:"
    ws['B6'] = invoice.status.upper()
    
    # Items header
    headers = ['Description', 'Quantity', 'Unit Price', 'Amount']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=8, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    
    # Items
    row = 9
    for item in items:
        ws.cell(row=row, column=1, value=item.description)
        ws.cell(row=row, column=2, value=item.quantity).number_format = '0.00'
        ws.cell(row=row, column=3, value=float(item.unit_price)).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=float(item.amount)).number_format = '#,##0.00'
        row += 1
    
    # Totals
    row += 1
    ws.cell(row=row, column=3, value="Subtotal:").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(invoice.subtotal)).number_format = '#,##0.00'
    
    row += 1
    ws.cell(row=row, column=3, value=f"Tax ({invoice.tax_rate}%):").font = Font(bold=True)
    ws.cell(row=row, column=4, value=float(invoice.tax_amount)).number_format = '#,##0.00'
    
    row += 1
    ws.cell(row=row, column=3, value="TOTAL:").font = Font(bold=True, size=12)
    total_cell = ws.cell(row=row, column=4, value=float(invoice.total_amount))
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True, size=12)
    total_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Adjust columns
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    
    # Save
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f'invoice_{invoice.invoice_number}.xlsx'
    return output, filename


def create_invoices_list_excel(invoices):
    """Create Excel file for invoice list
    
    Args:
        invoices: List of Invoice objects
        
    Returns:
        tuple: (BytesIO object with Excel file, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoices"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Invoice Number', 'Client Name', 'Project', 'Issue Date', 'Due Date',
        'Status', 'Payment Status', 'Subtotal', 'Tax Rate (%)', 'Tax Amount',
        'Total Amount', 'Amount Paid', 'Outstanding', 'Currency', 'Created By', 'Created At'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, invoice in enumerate(invoices, 2):
        data = [
            invoice.invoice_number,
            invoice.client_name or 'N/A',
            invoice.project.name if invoice.project else 'N/A',
            invoice.issue_date.strftime('%Y-%m-%d') if invoice.issue_date else '',
            invoice.due_date.strftime('%Y-%m-%d') if invoice.due_date else '',
            invoice.status or 'draft',
            invoice.payment_status or 'unpaid',
            float(invoice.subtotal or 0),
            float(invoice.tax_rate or 0),
            float(invoice.tax_amount or 0),
            float(invoice.total_amount or 0),
            float(invoice.amount_paid or 0),
            float(invoice.outstanding_amount or 0),
            invoice.currency_code or 'USD',
            invoice.creator.display_name if invoice.creator else 'Unknown',
            invoice.created_at.strftime('%Y-%m-%d %H:%M') if invoice.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Format number columns
            if col_num in [8, 10, 11, 12, 13]:  # Money columns
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
            elif col_num == 9:  # Tax rate percentage
                if isinstance(value, (int, float)):
                    cell.number_format = '0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary at the bottom
    last_row = len(invoices) + 2
    ws.cell(row=last_row + 1, column=1, value="Summary")
    ws.cell(row=last_row + 1, column=1).font = Font(bold=True)
    
    total_invoiced = sum(float(inv.total_amount or 0) for inv in invoices)
    total_paid = sum(float(inv.amount_paid or 0) for inv in invoices)
    total_outstanding = sum(float(inv.outstanding_amount or 0) for inv in invoices)
    
    ws.cell(row=last_row + 2, column=1, value="Total Invoiced:")
    ws.cell(row=last_row + 2, column=2, value=total_invoiced).number_format = '#,##0.00'
    ws.cell(row=last_row + 3, column=1, value="Total Paid:")
    ws.cell(row=last_row + 3, column=2, value=total_paid).number_format = '#,##0.00'
    ws.cell(row=last_row + 4, column=1, value="Total Outstanding:")
    ws.cell(row=last_row + 4, column=2, value=total_outstanding).number_format = '#,##0.00'
    ws.cell(row=last_row + 5, column=1, value="Total Invoices:")
    ws.cell(row=last_row + 5, column=2, value=len(invoices))
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'invoices_list_{timestamp}.xlsx'
    
    return output, filename


def create_payments_list_excel(payments):
    """Create Excel file for payment list
    
    Args:
        payments: List of Payment objects
        
    Returns:
        tuple: (BytesIO object with Excel file, filename)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Payments"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Payment ID', 'Invoice Number', 'Client Name', 'Amount', 'Currency',
        'Gateway Fee', 'Net Amount', 'Payment Date', 'Method', 'Reference',
        'Status', 'Received By', 'Gateway Transaction ID', 'Notes', 'Created At'
    ]
    
    # Write headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_num, payment in enumerate(payments, 2):
        data = [
            payment.id,
            payment.invoice.invoice_number if payment.invoice else 'N/A',
            payment.invoice.client_name if payment.invoice else 'N/A',
            float(payment.amount or 0),
            payment.currency or 'EUR',
            float(payment.gateway_fee or 0),
            float(payment.net_amount or payment.amount or 0),
            payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
            payment.method or 'N/A',
            payment.reference or '',
            payment.status or 'completed',
            payment.receiver.display_name if payment.receiver else 'N/A',
            payment.gateway_transaction_id or '',
            payment.notes or '',
            payment.created_at.strftime('%Y-%m-%d %H:%M') if payment.created_at else ''
        ]
        
        for col_num, value in enumerate(data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border
            
            # Format number columns
            if col_num in [4, 6, 7]:  # Money columns
                if isinstance(value, (int, float)):
                    cell.number_format = '#,##0.00'
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column].width = adjusted_width
    
    # Add summary at the bottom
    last_row = len(payments) + 2
    ws.cell(row=last_row + 1, column=1, value="Summary")
    ws.cell(row=last_row + 1, column=1).font = Font(bold=True)
    
    total_amount = sum(float(p.amount or 0) for p in payments)
    total_fees = sum(float(p.gateway_fee or 0) for p in payments if p.gateway_fee)
    total_net = sum(float(p.net_amount or p.amount or 0) for p in payments)
    completed_count = sum(1 for p in payments if p.status == 'completed')
    
    ws.cell(row=last_row + 2, column=1, value="Total Amount:")
    ws.cell(row=last_row + 2, column=2, value=total_amount).number_format = '#,##0.00'
    ws.cell(row=last_row + 3, column=1, value="Total Gateway Fees:")
    ws.cell(row=last_row + 3, column=2, value=total_fees).number_format = '#,##0.00'
    ws.cell(row=last_row + 4, column=1, value="Total Net Amount:")
    ws.cell(row=last_row + 4, column=2, value=total_net).number_format = '#,##0.00'
    ws.cell(row=last_row + 5, column=1, value="Total Payments:")
    ws.cell(row=last_row + 5, column=2, value=len(payments))
    ws.cell(row=last_row + 6, column=1, value="Completed Payments:")
    ws.cell(row=last_row + 6, column=2, value=completed_count)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'payments_list_{timestamp}.xlsx'
    
    return output, filename

